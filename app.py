import pandas as pd
import os
import re
import io
import base64
import plotly.express as px
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- APP START ---
st.set_page_config(page_title="Recon Tool", layout="wide")

# --- SECURITY GATEKEEPER ---
def check_password():
    if "password_correct" not in st.session_state:
        st.session_state["password_correct"] = False
    if st.session_state["password_correct"]:
        return True

    col1, col2, col3 = st.columns([1,2,1])
    with col2:
        st.title("Secure Access")
        user_pwd = st.text_input("Institutional Password", type="password")
        if st.button("Unlock System"):
            # FIX: Ensure we check the secret correctly without crashing
            if "INSTITUTIONAL_PASSWORD" in st.secrets and user_pwd == st.secrets["INSTITUTIONAL_PASSWORD"]: 
                st.session_state["password_correct"] = True
                st.rerun()
            else:
                st.error("🚫 Access Denied")
    return False

# This is the "Gate": If not logged in, the rest of the code doesn't execute
if check_password():

    # --- STYLING CONSTANTS ---
    NAVY_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    GREY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    WHITE_TEXT = Font(color="FFFFFF", bold=True)
    BLACK_BOLD = Font(color="000000", bold=True)
    RED_BOLD = Font(color="CC0000", bold=True)
    DOUBLE_BORDER = Border(top=Side(style='double'), bottom=Side(style='double'))
    THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    st.title("📂 Institutional Reconciliation System")

    # --- SIDEBAR TOOLS ---
    st.sidebar.header("🔍 Quick Search")
    search_ref = st.sidebar.text_input("Enter FCM Reference to Track")

    col1, col2 = st.columns(2)
    with col1:
        gl_uploads = st.file_uploader("Upload GL Excel Files", type=['xlsx', 'xls'], accept_multiple_files=True)
    with col2:
        csv_uploads = st.file_uploader("Upload NIBSS CSV Files", type=['csv'], accept_multiple_files=True)

    def heavy_clean_file(uploaded_file):
        try:
            # FIX: reset pointer so the file can be read even if Streamlit peaked at it
            uploaded_file.seek(0)
            content = uploaded_file.read()
            text = content.replace(b'\x00', b'').decode('utf-8', errors='ignore')
            if '\t' in text: text = text.replace(',', ';').replace('\t', ',')
            return io.StringIO(text)
        except: return None

    def clean_num(col):
        col_clean = col.astype(str).str.replace(r'[^-0-9.]', '', regex=True)
        return pd.to_numeric(col_clean, errors='coerce').fillna(0)

    if st.button("🚀 Run Reconciliation"):
        if gl_uploads and csv_uploads:
            run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            
            # --- DATA LOADING & PROCESSING ---
            all_gl, all_csv, csv_dict = [], [], {}
            for f in gl_uploads:
                df = pd.read_excel(f)
                df.columns = [str(c).strip() for c in df.columns]
                if 'Description' not in df.columns:
                    st.error(f"❌ Error: File '{f.name}' missing 'Description'.")
                    st.stop()
                for c in ['Deposit', 'Withdrawal', 'Balance']:
                    if c in df.columns: df[c] = clean_num(df[c])
                all_gl.append(df)
            df_gl_input = pd.concat(all_gl, ignore_index=True) if all_gl else pd.DataFrame()

            for f in csv_uploads:
                buffer = heavy_clean_file(f)
                if buffer:
                    df = pd.read_csv(buffer, on_bad_lines='skip', engine='python')
                    df.columns = [str(c).strip().lower() for c in df.columns]
                    df.rename(columns={'uniquereference': 'unique_reference', 'remittedamount': 'remitted_amount', 'mdaname': 'mda_name'}, inplace=True)
                    for c in ['remitted_amount', 'collected_amount', 'fee']:
                        if c in df.columns: df[c] = clean_num(df[c])
                    df['source_file'] = f.name
                    if 'mda_name' in df.columns:
                        df['mda_name'] = df['mda_name'].astype(str).str.strip().str.upper()
                    all_csv.append(df)
                    csv_dict[f.name] = df
            df_csv_input = pd.concat(all_csv, ignore_index=True) if all_csv else pd.DataFrame()

            gl_review = df_gl_input.copy() if not df_gl_input.empty else pd.DataFrame(columns=['Description', 'Deposit', 'Withdrawal', 'Reference'])
            
            # GL_Reference logic
            if 'Reference' in gl_review.columns:
                gl_review['GL_Reference'] = gl_review['Reference'].astype(str).replace('nan', '')
            else:
                gl_review['GL_Reference'] = ""

            if 'Description' in gl_review.columns:
                gl_review['Reference'] = gl_review['Description'].astype(str).str.extract(r'(FCM\d{17})', expand=False).fillna("")
            else:
                gl_review['Reference'] = ""
            
            cols = list(gl_review.columns)
            if 'GL_Reference' in cols:
                cols.insert(2, cols.pop(cols.index('GL_Reference')))
                gl_review = gl_review[cols]
                
            csv_totals = df_csv_input.groupby('unique_reference')['remitted_amount'].sum().to_dict() if not df_csv_input.empty else {}
            gl_review['NIBSS_remitted'] = gl_review['Reference'].map(csv_totals).fillna(0)
            gl_review['NIBSS_reference'] = gl_review.apply(lambda x: x['Reference'] if x['Reference'] in csv_totals and x['Reference'] != "" else "", axis=1)
            gl_review['Variance'] = gl_review['NIBSS_remitted'] - gl_review['Deposit']

            gl_match_map = gl_review.groupby('Reference')['Deposit'].sum().to_dict()
            nibss_review = df_csv_input.copy() if not df_csv_input.empty else pd.DataFrame(columns=['unique_reference', 'remitted_amount', 'bank_id'])

            if not nibss_review.empty:
                b_idx = list(nibss_review.columns).index('bank_id') + 1 if 'bank_id' in nibss_review.columns else 1
                nibss_review.insert(b_idx, 'Kachasi_ref', nibss_review['unique_reference'].map(lambda x: x if x in gl_match_map else ""))
                nibss_review.insert(b_idx+1, 'Kachasi_In_GL', nibss_review['unique_reference'].map(gl_match_map).fillna(0))
                nibss_review.insert(b_idx+2, 'Settle_Variance', nibss_review['remitted_amount'] - nibss_review['Kachasi_In_GL'])

            # --- CALCULATIONS ---
            matched_mask_gl = (gl_review['NIBSS_reference'] != "")
            total_matched_gl_dep = gl_review[matched_mask_gl]['Deposit'].sum()
            total_matched_gl_nibss = gl_review[matched_mask_gl]['NIBSS_remitted'].sum()
            unmatched_gl_dep = gl_review[~matched_mask_gl & (gl_review['Deposit'] > 0)]['Deposit'].sum()
            bridging_diff = total_matched_gl_dep - total_matched_gl_nibss
            csv_vs_kachasi_diff = total_matched_gl_nibss - total_matched_gl_dep
            
            is_m = (nibss_review['Kachasi_ref'] != "") if 'Kachasi_ref' in nibss_review.columns else pd.Series([False]*len(nibss_review))
            # FIX: Ensure mda_name exists before checking
            mda_col = nibss_review['mda_name'] if 'mda_name' in nibss_review.columns else pd.Series([""]*len(nibss_review))
            is_c = ((nibss_review['Kachasi_ref'] == "") & (mda_col == 'NIGERIA CUSTOM SERVICES') & (nibss_review['unique_reference'].str.startswith('F', na=False)))
            
            unmatched_csv_customs = nibss_review[is_c]['remitted_amount'].sum() if not nibss_review.empty else 0
            unmatched_csv_others = nibss_review[~is_m & ~is_c]['remitted_amount'].sum() if not nibss_review.empty else 0

            # --- BROWSER DASHBOARD ---
            st.markdown("---")
            st.subheader("📊 Executive Insights")
            
            v1, v2 = st.columns([1, 1.5])
            with v1:
                if 'mda_name' in df_csv_input.columns:
                    mda_share = df_csv_input.groupby('mda_name')['remitted_amount'].sum().reset_index()
                    mda_share = mda_share[mda_share['remitted_amount'] > 0]
                    if not mda_share.empty:
                        fig = px.pie(mda_share, values='remitted_amount', names='mda_name', hole=0.4, title="Remittance Market Share")
                        fig.update_layout(showlegend=False, height=350)
                        st.plotly_chart(fig, use_container_width=True)

            with v2:
                m1, m2, m3 = st.columns(3)
                m1.metric("Overall GL Deposit", f"₦{gl_review['Deposit'].sum():,.2f}")
                m2.metric("Overall NIBSS Remittance", f"₦{df_csv_input['remitted_amount'].sum():,.2f}")
                m3.metric("Net System Variance", f"₦{gl_review['Variance'].sum():,.2f}")

            st.markdown("### Detailed Reconciliation Breakdown")
            d1, d2, d3 = st.columns(3)
            with d1:
                st.write("**Bridging (Excel to CSV)**")
                df_bridge = pd.DataFrame({"Description": ["Matched GL Deposit", "Matched NIBSS Remitted", "Difference (Matched GL vs NIBSS)", "Unmatched GL Dep"], "Value": [total_matched_gl_dep, total_matched_gl_nibss, bridging_diff, unmatched_gl_dep]}).set_index("Description")
                st.table(df_bridge.style.format("₦{:,.2f}"))
            with d2:
                st.write("**CSV to Excel Analysis**")
                df_comp = pd.DataFrame({"Description": ["Total Matched CSV Remittance", "Total Matched Kachasi Credit", "Difference (CSV vs Kachasi)"], "Value": [total_matched_gl_nibss, total_matched_gl_dep, csv_vs_kachasi_diff]}).set_index("Description")
                st.table(df_comp.style.format("₦{:,.2f}"))
            with d3:
                st.write("**Unmatched CSV Categorization**")
                st.table(pd.DataFrame({"Description": ["Customs (NCS)", "Other MDAs", "Total Unmatched CSV"], "Value": [unmatched_csv_customs, unmatched_csv_others, unmatched_csv_customs + unmatched_csv_others]}).set_index("Description").style.format("₦{:,.2f}"))

            # --- EXCEL OUTPUT GENERATION ---
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Summary logic
                summary_sections = [
                    ['EXECUTIVE RECONCILIATION DASHBOARD', ''],
                    ['Run Timestamp', run_time],
                    ['Overall GL Deposit', gl_review['Deposit'].sum()],
                    ['Overall NIBSS Remittance', df_csv_input['remitted_amount'].sum()],
                    ['Net System Variance', gl_review['Variance'].sum()],
                    ['', ''],
                    ['--- EXCEL TO CSV ANALYSIS (Bridging) ---', 'VALUE'],
                    ['Total Matched Excel Deposit', total_matched_gl_dep],
                    ['Total Matched NIBSS Remitted', total_matched_gl_nibss],
                    ['Difference (Matched Excel vs NIBSS)', bridging_diff],
                    ['Total Unmatched Excel Deposit (Exceptions)', unmatched_gl_dep],
                    ['Matched Item Count', matched_mask_gl.sum()],
                    ['', ''],
                    ['--- CSV TO EXCEL ANALYSIS (Categorized) ---', 'VALUE'],
                    ['Total Matched CSV Remittance', total_matched_gl_nibss],
                    ['Total Matched Kachasi Credit', total_matched_gl_dep],
                    ['Difference (Matched CSV vs Kachasi)', csv_vs_kachasi_diff],
                    ['', ''],
                    ['Total Unmatched CSV - NIGERIA CUSTOM SERVICES', unmatched_csv_customs],
                    ['Total Unmatched CSV - OTHER MDAs', unmatched_csv_others],
                    ['Total Unmatched CSV Remittance', unmatched_csv_customs + unmatched_csv_others]
                ]
                pd.DataFrame(summary_sections).to_excel(writer, sheet_name='Executive Summary', index=False, header=False)
                ws_sum = writer.sheets['Executive Summary']
                for i, row in enumerate(summary_sections, 1):
                    ws_sum[f'A{i}'].border = ws_sum[f'B{i}'].border = THIN_BORDER
                    if i == 1 or '---' in str(row[0]):
                        ws_sum[f'A{i}'].fill = NAVY_FILL; ws_sum[f'A{i}'].font = WHITE_TEXT; ws_sum[f'B{i}'].fill = NAVY_FILL
                    if 'Difference' in str(row[0]):
                        ws_sum[f'A{i}'].font = RED_BOLD; ws_sum[f'B{i}'].font = RED_BOLD
                    if isinstance(row[1], (int, float)): ws_sum[f'B{i}'].number_format = '#,##0.00'

                def write_block(ws, df, start_row, label, sum_cols):
                    if df.empty: return start_row
                    df_clean = df.fillna('')
                    df_clean.to_excel(writer, sheet_name=ws.title, startrow=start_row, index=False)
                    h_row = start_row + 1 
                    for r in range(h_row, h_row + len(df_clean) + 1):
                        for c in range(1, len(df_clean.columns) + 1):
                            cell = ws.cell(row=r, column=c)
                            cell.border = THIN_BORDER
                            if r == h_row:
                                cell.fill = NAVY_FILL; cell.font = WHITE_TEXT; cell.alignment = Alignment(horizontal='center')
                            else:
                                if r % 2 == 0: cell.fill = GREY_FILL
                                col_name = str(df_clean.columns[c-1]).lower()
                                if any(x in col_name for x in ['amount', 'fee', 'deposit', 'withdrawal', 'variance', 'remitted', 'in_gl']):
                                    cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal='right')
                    t_row = h_row + len(df_clean) + 1
                    ws.cell(row=t_row, column=2, value=f"SUB-TOTAL: {label}").font = BLACK_BOLD
                    for i, col in enumerate(df_clean.columns, 1):
                        if col in sum_cols:
                            cell = ws.cell(row=t_row, column=i, value=df[col].sum())
                            cell.font = BLACK_BOLD; cell.number_format = '#,##0.00'
                    return t_row + 2

                def write_grand_total(ws, df, row, label, sum_cols):
                    ws.cell(row=row, column=2, value=label).font = RED_BOLD
                    for i, col in enumerate(df.columns, 1):
                        if col in sum_cols:
                            cell = ws.cell(row=row, column=i, value=df[col].sum())
                            cell.font = BLACK_BOLD; cell.border = DOUBLE_BORDER; cell.number_format = '#,##0.00'
                    return row + 2

                ws_gl = writer.book.create_sheet('GL_Review')
                gl_sums = ['Deposit', 'Withdrawal', 'NIBSS_remitted', 'Variance']
                r = write_block(ws_gl, gl_review[gl_review['NIBSS_reference'] != ""], 0, "MATCHED GL", gl_sums)
                r = write_block(ws_gl, gl_review[(gl_review['NIBSS_reference'] == "") & (gl_review['Deposit'] > 0)], r, "UNMATCHED DEPOSIT", gl_sums)
                r = write_block(ws_gl, gl_review[(gl_review['NIBSS_reference'] == "") & (gl_review['Withdrawal'] > 0)], r, "UNMATCHED WITHDRAWAL", gl_sums)
                write_grand_total(ws_gl, gl_review, r, "GRAND TOTAL (GL_REVIEW)", gl_sums)

                ws_nr = writer.book.create_sheet('NIBSS_Review')
                nr_sums = ['remitted_amount', 'collected_amount', 'fee', 'Kachasi_In_GL', 'Settle_Variance']
                r_n = write_block(ws_nr, nibss_review[is_m], 0, "MATCHED NIBSS", nr_sums)
                r_n = write_block(ws_nr, nibss_review[is_c], r_n, "UNMATCHED CUSTOMS (F-REF)", nr_sums)
                r_n = write_block(ws_nr, nibss_review[(~is_m) & (~is_c)], r_n, "OTHER UNMATCHED", nr_sums)
                write_grand_total(ws_nr, nibss_review, r_n, "GRAND TOTAL (NIBSS_REVIEW)", nr_sums)

                mda_configs = [
                    ('NIGERIA CUSTOM SERVICES', 'Customs_Extract'), 
                    ('FEDERAL MINISTRY OF FINANCE, BUDGET AND NATIONAL PLANNING - HQTRS', 'NESS_Extract'),
                    ('OFFICE OF THE CHIEF SECURITY OFFICER TO THE PRESIDENT', 'CyberSec_Extract')
                ]
                for mda_target, sname in mda_configs:
                    ws_ex = writer.book.create_sheet(sname)
                    ptr, pool = 0, []
                    for fn, df_f in csv_dict.items():
                        if 'mda_name' in df_f.columns and 'unique_reference' in df_f.columns:
                            cond = (df_f['mda_name'] == mda_target) & (df_f['unique_reference'].str.startswith('F', na=False))
                            ext = df_f[cond]
                            if not ext.empty:
                                ws_ex.cell(row=ptr+1, column=1, value=f"FILE: {fn}").font = BLACK_BOLD
                                ptr = write_block(ws_ex, ext, ptr + 1, fn, ['remitted_amount', 'collected_amount', 'fee'])
                                pool.append(ext)
                    if not pool:
                        zero_df = pd.DataFrame([["NO RECORDS FOUND", 0.00, 0.00, 0.00, mda_target]], columns=['unique_reference', 'remitted_amount', 'collected_amount', 'fee', 'mda_name'])
                        write_block(ws_ex, zero_df, 0, "ZERO EXTRACT", ['remitted_amount', 'collected_amount', 'fee'])
                    else: 
                        write_grand_total(ws_ex, pd.concat(pool), ptr, f"GRAND TOTAL ({sname})", ['remitted_amount', 'collected_amount', 'fee'])

                ws_log = writer.book.create_sheet('Run_Log')
                log_data = [['RECONCILIATION AUDIT LOG', ''], ['Processed At:', run_time], ['', ''], ['SOURCE FILES USED:', 'TYPE']]
                for f in gl_uploads: log_data.append([f.name, 'EXCEL / GL'])
                for f in csv_uploads: log_data.append([f.name, 'CSV / NIBSS'])
                pd.DataFrame(log_data).to_excel(writer, sheet_name='Run_Log', index=False, header=False)

                for sheet in writer.book.worksheets:
                    for col in sheet.columns: sheet.column_dimensions[col[0].column_letter].width = 28

            st.download_button(label="📥 Download Executive Report", data=output.getvalue(), file_name="Executive_Recon_Report.xlsx")
        else:
            st.error("Please upload both GL and NIBSS files.")
